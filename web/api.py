"""Structured chat API for CheetahClaws web UI.

Bridges the synchronous agent.run() generator to WebSocket event streaming,
following the same pattern as the Telegram/Slack/WeChat bridges:
wire RuntimeContext callbacks → run agent on background thread → push events.
"""
from __future__ import annotations

import base64
import copy
import csv
import io
import json
import os
import queue
import sys
import threading
import time
import traceback
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# Ensure the package root is importable (web/ is a subpackage)
_PKG_ROOT = str(Path(__file__).resolve().parent.parent)
if _PKG_ROOT not in sys.path:
    sys.path.insert(0, _PKG_ROOT)


# ── Event envelope ─────────────────────────────────────────────────────────

@dataclass
class ChatEvent:
    """JSON-serializable event sent to browser via WebSocket."""
    type: str       # text_chunk | thinking_chunk | tool_start | tool_end |
                    # permission_request | permission_response | turn_done |
                    # error | status
    data: dict = field(default_factory=dict)
    ts: float = field(default_factory=time.time)

    def to_json(self) -> str:
        return json.dumps({"type": self.type, "data": self.data, "ts": self.ts})


# ── Slash command handler (can't import from cheetahclaws.py — it has
#    top-level code that runs on import).  Build our own from commands/*.
# ──────────────────────────────────────────────────────────────────────────

_WEB_COMMANDS: dict | None = None


def _get_web_commands() -> dict:
    """Lazily build the slash command registry from commands/ submodules."""
    global _WEB_COMMANDS
    if _WEB_COMMANDS is not None:
        return _WEB_COMMANDS

    cmds: dict = {}
    # Import each group separately so partial failures don't block others
    _imports = [
        # (module, [(cmd_name, func_name), ...])
        ("commands.core", [
            ("help", "cmd_help"), ("clear", "cmd_clear"),
            ("context", "cmd_context"), ("cost", "cmd_cost"),
            ("compact", "cmd_compact"), ("status", "cmd_status"),
            ("export", "cmd_export"), ("copy", "cmd_copy"),
            ("doctor", "cmd_doctor"), ("init", "cmd_init"),
            ("proactive", "cmd_proactive"), ("image", "cmd_image"),
            ("img", "cmd_image"),
        ]),
        ("commands.session", [
            ("save", "cmd_save"), ("load", "cmd_load"),
            ("resume", "cmd_resume"), ("search", "cmd_search"),
            ("history", "cmd_history"), ("cloudsave", "cmd_cloudsave"),
            ("exit", "cmd_exit"), ("quit", "cmd_exit"),
        ]),
        ("commands.config_cmd", [
            ("model", "cmd_model"), ("config", "cmd_config"),
            ("verbose", "cmd_verbose"), ("thinking", "cmd_thinking"),
            ("permissions", "cmd_permissions"), ("cwd", "cmd_cwd"),
        ]),
        ("commands.advanced", [
            ("brainstorm", "cmd_brainstorm"), ("worker", "cmd_worker"),
            ("ssj", "cmd_ssj"), ("skills", "cmd_skills"),
            ("memory", "cmd_memory"), ("agents", "cmd_agents"),
            ("mcp", "cmd_mcp"), ("plugin", "cmd_plugin"),
            ("tasks", "cmd_tasks"), ("task", "cmd_tasks"),
        ]),
        ("commands.checkpoint_plan", [
            ("plan", "cmd_plan"), ("checkpoint", "cmd_checkpoint"),
        ]),
        ("commands.agent_cmd", [
            ("agent", "cmd_agent"),
        ]),
        ("commands.monitor_cmd", [
            ("subscribe", "cmd_subscribe"),
            ("subscriptions", "cmd_subscriptions"),
            ("subs", "cmd_subscriptions"),
            ("unsubscribe", "cmd_unsubscribe"),
            ("monitor", "cmd_monitor"),
        ]),
        # External bridges — telegram / slack / wechat / voice.
        # Each lives in its own module so missing deps (e.g. sounddevice for
        # voice) just skip that one command instead of blocking the rest.
        ("bridges.telegram", [("telegram", "cmd_telegram")]),
        ("bridges.slack",    [("slack",    "cmd_slack")]),
        ("bridges.wechat",   [("wechat",   "cmd_wechat"),
                              ("weixin",   "cmd_wechat")]),
        ("modular.voice.cmd", [("voice",   "cmd_voice")]),
    ]
    import importlib
    for mod_name, pairs in _imports:
        try:
            mod = importlib.import_module(mod_name)
            for cmd_name, func_name in pairs:
                fn = getattr(mod, func_name, None)
                if fn:
                    cmds[cmd_name] = fn
        except ImportError:
            pass
    _WEB_COMMANDS = cmds
    return cmds


def _web_handle_slash(line: str, state, config):
    """Handle /command. Returns True if handled, or sentinel tuple."""
    if not line.startswith("/"):
        return False
    parts = line[1:].split(None, 1)
    if not parts:
        return False
    cmd = parts[0].lower()
    args = parts[1] if len(parts) > 1 else ""
    commands = _get_web_commands()
    handler = commands.get(cmd)
    if handler:
        result = handler(args, state, config)
        # Sentinel tuples need special handling by the caller
        _SENTINELS = ("__voice__", "__image__", "__brainstorm__", "__worker__",
                      "__ssj_cmd__", "__ssj_query__", "__ssj_debate__",
                      "__ssj_passthrough__", "__ssj_promote_worker__", "__plan__")
        if isinstance(result, tuple) and result[0] in _SENTINELS:
            return result
        return True
    print(f"Unknown command: /{cmd}  (type /help for commands)")
    return True


# ── Chat Session ───────────────────────────────────────────────────────────

_IDLE_TIMEOUT = 1800  # 30 min before session is considered stale

_SAFE_CONFIG_KEYS = frozenset({
    "model", "permission_mode", "max_tokens", "verbose", "thinking",
    "thinking_budget", "max_tool_output", "max_agent_depth",
    "shell_policy", "log_level",
})

_WRITABLE_CONFIG_KEYS = frozenset({
    "model", "permission_mode", "verbose", "thinking",
    "thinking_budget", "max_tokens",
    # API keys — written to session config only, not persisted to disk
    "anthropic_api_key", "openai_api_key", "gemini_api_key",
    "kimi_api_key", "qwen_api_key", "zhipu_api_key",
    "deepseek_api_key", "minimax_api_key", "custom_api_key",
    "custom_base_url", "ollama_base_url",
})

# Keys that contain secrets — never expose in GET responses
_SECRET_KEYS = frozenset({
    "anthropic_api_key", "openai_api_key", "gemini_api_key",
    "kimi_api_key", "qwen_api_key", "zhipu_api_key",
    "deepseek_api_key", "minimax_api_key", "custom_api_key",
})

_API_KEY_CONFIG_MAP = {
    "anthropic": "anthropic_api_key",
    "openai": "openai_api_key",
    "gemini": "gemini_api_key",
    "kimi": "kimi_api_key",
    "qwen": "qwen_api_key",
    "zhipu": "zhipu_api_key",
    "deepseek": "deepseek_api_key",
    "minimax": "minimax_api_key",
    "custom": "custom_api_key",
}

_MAX_ATTACHMENT_BYTES = 20 * 1024 * 1024
_MAX_ATTACHMENT_TEXT_CHARS = 80000
_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
_SPREADSHEET_EXTS = {".xlsx", ".xlsm", ".xls"}
_TEXT_EXTS = {".txt", ".md", ".csv", ".tsv", ".json", ".log", ".xml", ".yaml", ".yml"}


def _decode_text_attachment(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _spreadsheet_attachment_summary(raw: bytes, ext: str) -> str:
    chunks: list[str] = []
    if ext in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "[附件解析失败：当前环境未安装 openpyxl，无法读取 .xlsx/.xlsm 表格。请安装 openpyxl，或让用户另存为 .csv 后重新上传。不要用系统已有交易记录替代这个附件内容。]"
        try:
            workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        except Exception:
            return "[附件解析失败：该 .xlsx/.xlsm 文件不是可识别的 Excel 表格或文件已损坏。请另存为 .xlsx / .csv 后重新上传。不要用系统已有交易记录替代这个附件内容。]"
        for sheet in workbook.worksheets[:5]:
            chunks.append(f"# Sheet: {sheet.title}")
            for row_index, row in enumerate(sheet.iter_rows(max_row=300, max_col=30, values_only=True), start=1):
                values = ["" if cell is None else str(cell) for cell in row]
                if any(value.strip() for value in values):
                    chunks.append(f"{row_index}\t" + "\t".join(values))
    elif ext == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError:
            return "[附件解析失败：这是 .xls 老版 Excel 文件，但当前环境未安装 xlrd，无法读取表格内容。请安装 xlrd，或让用户另存为 .xlsx / .csv 后重新上传。不要用系统已有交易记录替代这个附件内容。]"
        try:
            workbook = xlrd.open_workbook(file_contents=raw)
        except Exception as exc:
            return f"[附件解析失败：该 .xls 文件无法被 xlrd 读取（{exc}）。请另存为 .xlsx / .csv 后重新上传。不要用系统已有交易记录替代这个附件内容。]"
        for sheet in workbook.sheets()[:5]:
            chunks.append(f"# Sheet: {sheet.name}")
            for row_index in range(min(sheet.nrows, 300)):
                values = ["" if sheet.cell_value(row_index, col) in (None, "") else str(sheet.cell_value(row_index, col)) for col in range(min(sheet.ncols, 30))]
                if any(value.strip() for value in values):
                    chunks.append(f"{row_index + 1}\t" + "\t".join(values))
    return "\n".join(chunks) or "[附件解析失败：表格中没有发现可读取的非空单元格。不要用系统已有交易记录替代这个附件内容。]"


def _xlsx_attachment_summary(raw: bytes) -> str:
    return _spreadsheet_attachment_summary(raw, ".xlsx")


def _csv_attachment_summary(raw: bytes, ext: str) -> str:
    text = _decode_text_attachment(raw)
    if ext not in {".csv", ".tsv"}:
        return text
    delimiter = "\t" if ext == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = []
    for index, row in enumerate(reader):
        if index >= 80:
            rows.append("...[truncated]")
            break
        rows.append("\t".join(row))
    return "\n".join(rows)


def _pdf_attachment_summary(raw: bytes) -> str:
    try:
        import fitz
    except ImportError:
        return "[PDF attachment received, but pymupdf is not installed. Install cheetahclaws[files] to extract PDF text.]"
    chunks: list[str] = []
    with fitz.open(stream=raw, filetype="pdf") as document:
        for page_index, page in enumerate(document[:10], start=1):
            text = page.get_text("text").strip()
            if text:
                chunks.append(f"# Page {page_index}\n{text}")
    return "\n\n".join(chunks) or "[PDF attachment received, but no extractable text was found.]"


def _normalize_image_mime(mime: str, ext: str) -> str:
    value = (mime or "").lower()
    allowed = {"image/jpeg", "image/png", "image/gif", "image/webp"}
    if value in allowed:
        return value
    if ext in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if ext == ".png":
        return "image/png"
    if ext == ".gif":
        return "image/gif"
    if ext == ".webp":
        return "image/webp"
    return "image/png"


def _attachment_to_context(name: str, mime: str, raw: bytes) -> tuple[str, dict | None]:
    ext = Path(name or "attachment").suffix.lower()
    if ext in _IMAGE_EXTS or mime.startswith("image/"):
        image_mime = _normalize_image_mime(mime, ext)
        return (
            f"[Image attachment: {name or 'image'} | {image_mime}]",
            {"data": base64.b64encode(raw).decode("utf-8"), "mime": image_mime},
        )
    if ext in _SPREADSHEET_EXTS:
        return _spreadsheet_attachment_summary(raw, ext), None
    if ext == ".pdf":
        return _pdf_attachment_summary(raw), None
    if ext in _TEXT_EXTS or mime.startswith("text/"):
        return _csv_attachment_summary(raw, ext), None
    return f"[Attachment received: {name or 'file'} ({mime or 'unknown type'}, {len(raw)} bytes). Binary content was not expanded into the prompt.]", None


class ChatSession:
    """One agent conversation, bridged to WebSocket clients.

    Persistence: session metadata and message history are mirrored to SQLite
    via web.db.repo. The in-memory `messages` list is a write-through cache
    for fast replay; event queue/buffer stay in-memory only.
    """

    def __init__(self, base_config: dict, user_id: int, *,
                 session_id: Optional[str] = None,
                 title: Optional[str] = None):
        from web import db as _db
        _db.init_db()

        # Hydrate-from-DB path vs new-session path
        existing = (_db.repo.get_session(session_id, user_id)
                    if session_id else None)

        self.session_id: str = (existing["id"] if existing
                                else (session_id or uuid.uuid4().hex[:12]))
        self.user_id: int = user_id
        self.title: str = (existing["title"] if existing
                           else (title or "New chat"))
        self.created_at: float = (existing["created_at"] if existing
                                  else time.time())
        self.last_active: float = time.time()

        # Deep-copy config so permission_mode changes don't leak
        base = copy.deepcopy(base_config)
        if existing and existing.get("config"):
            base.update(existing["config"])
        self.config: dict = base
        if self.config.get("rich_business_mode"):
            from tools.rich_business import _RICH_BUSINESS_TOOLS
            self.config["allowed_tools"] = list(_RICH_BUSINESS_TOOLS)
            self.config["shell_policy"] = "deny"
            self.config["permission_mode"] = "auto"
        self.config["_session_id"] = self.session_id

        # Event fan-out: multiple WS clients can subscribe
        self._subscribers: list[queue.Queue] = []
        self._sub_lock = threading.Lock()

        # Buffer recent events so late-joining subscribers don't miss them.
        # Capped at 500 events; covers the gap between agent start and WS connect.
        self._event_buffer: list[ChatEvent] = []
        self._EVENT_BUFFER_MAX = 500

        # Agent state (in-process, NOT a PTY subprocess)
        self._agent_state = None  # type: ignore[assignment]
        self._agent_thread: Optional[threading.Thread] = None
        self._busy = threading.Event()

        # Message history for UI replay on reconnect (hydrated from DB)
        self.messages: list[dict] = (_db.repo.get_messages(self.session_id)
                                     if existing else [])
        self._msg_lock = threading.Lock()
        self._pending_attachments: list[dict] = []

        # Persist (create-or-update) metadata
        _db.repo.upsert_session(
            self.session_id, user_id,
            title=self.title,
            config={k: v for k, v in self.config.items()
                    if k in _SAFE_CONFIG_KEYS},
        )

        self._init_runtime()

    def _init_runtime(self):
        """Initialize RuntimeContext and AgentState."""
        from agent import AgentState
        import runtime

        self._agent_state = AgentState()
        ctx = runtime.get_session_ctx(self.session_id)
        ctx.agent_state = self._agent_state
        ctx.run_query = lambda msg: self.submit_prompt(msg)

    # ── Subscriber management ──────────────────────────────────────────

    def subscribe(self) -> queue.Queue:
        """Add a subscriber and replay any buffered events."""
        q: queue.Queue = queue.Queue(maxsize=2000)
        with self._sub_lock:
            # Replay buffered events so late-joiners don't miss anything
            for evt in self._event_buffer:
                try:
                    q.put_nowait(evt)
                except queue.Full:
                    break
            self._subscribers.append(q)
        return q

    def unsubscribe(self, q: queue.Queue):
        with self._sub_lock:
            try:
                self._subscribers.remove(q)
            except ValueError:
                pass

    def add_attachment(self, name: str, mime: str, data_b64: str) -> dict:
        if len(self._pending_attachments) >= 8:
            raise ValueError("Too many pending attachments")
        try:
            raw = base64.b64decode(data_b64, validate=True)
        except Exception as exc:
            raise ValueError("Invalid attachment encoding") from exc
        if not raw:
            raise ValueError("Attachment is empty")
        if len(raw) > _MAX_ATTACHMENT_BYTES:
            raise ValueError("Attachment is too large")
        context_text, image_part = _attachment_to_context(name, mime, raw)
        if len(context_text) > _MAX_ATTACHMENT_TEXT_CHARS:
            context_text = context_text[:_MAX_ATTACHMENT_TEXT_CHARS] + "\n...[truncated]"
        attachment = {
            "name": name or "attachment",
            "mime": mime or "application/octet-stream",
            "size": len(raw),
            "context": context_text,
            "image_part": image_part,
        }
        self._pending_attachments.append(attachment)
        self._broadcast(ChatEvent("attachment_added", {
            "name": attachment["name"],
            "mime": attachment["mime"],
            "size": attachment["size"],
            "is_image": bool(image_part),
        }))
        return {
            "name": attachment["name"],
            "mime": attachment["mime"],
            "size": attachment["size"],
            "is_image": bool(image_part),
        }

    def _consume_attachment_context(self) -> tuple[str, list[dict]]:
        if not self._pending_attachments:
            return "", []
        attachments = self._pending_attachments
        self._pending_attachments = []
        sections = [
            "\n\n以下是用户本轮主动上传的附件内容，不是服务器文件或代码仓库内容；"
            "你可以直接基于这些附件内容回答或调用 RICH 业务工具。"
            "如果附件内容显示解析失败、无法读取或被截断到不足以判断，必须直接说明无法查看该附件，"
            "不要查询或复述系统已有业务数据来冒充附件内容。"
        ]
        images = []
        for index, att in enumerate(attachments, start=1):
            sections.append(
                f"\n\n[Attachment {index}: {att['name']} | {att['mime']} | {att['size']} bytes]\n"
                f"{att['context']}"
            )
            if att.get("image_part"):
                images.append(att["image_part"])
        return "".join(sections), images

    def _broadcast(self, event: ChatEvent):
        with self._sub_lock:
            # Buffer for late-joining subscribers
            self._event_buffer.append(event)
            if len(self._event_buffer) > self._EVENT_BUFFER_MAX:
                self._event_buffer = self._event_buffer[-self._EVENT_BUFFER_MAX:]
            # Push to live subscribers
            for q in self._subscribers:
                try:
                    q.put_nowait(event)
                except queue.Full:
                    pass

    def _filter_text(self, text: str) -> str:
        """Apply output filter to sanitize agent text before broadcast."""
        if getattr(self, '_cached_filter', None) is None:
            from output_filter import create_filter
            self._cached_filter = create_filter(self.config)
        return self._cached_filter(text)

    # ── Prompt submission ──────────────────────────────────────────────

    def handle_slash_sync(self, line: str) -> list[dict]:
        """Handle a slash command synchronously. Returns list of event dicts
        to send back in the HTTP response.

        Synchronous events are returned via HTTP only — re-broadcasting them
        to WS subscribers would duplicate every reply in the chat UI, since
        the same client also iterates the HTTP `events` payload. Background
        threads spawned by the handler still broadcast normally, because
        `_broadcast` is restored before they emit anything.
        """
        events: list[dict] = []
        orig_broadcast = self._broadcast

        def capture_broadcast(event: ChatEvent):
            events.append({"type": event.type, "data": event.data})

        self._broadcast = capture_broadcast  # type: ignore
        try:
            self._handle_slash(line)
        finally:
            self._broadcast = orig_broadcast  # type: ignore
        return events

    def handle_slash_stream(self, line: str, event_callback):
        """Handle a slash command, calling event_callback(dict) for each event.
        Blocks until the command (including long-running ones) completes.
        Used by the SSE streaming endpoint.

        Events are delivered via the SSE callback only — re-broadcasting them
        to WS subscribers would duplicate every reply in the chat UI, since
        the same client also calls _handleEvent on the SSE stream.
        """
        done_event = threading.Event()
        orig_broadcast = self._broadcast

        def stream_broadcast(event: ChatEvent):
            event_callback({"type": event.type, "data": event.data})
            if event.type == "status" and event.data.get("state") == "idle":
                done_event.set()

        self._broadcast = stream_broadcast  # type: ignore
        try:
            self._handle_slash(line)
            # For long-running commands, wait until the bg thread finishes
            if self._busy.is_set():
                done_event.wait(timeout=600)  # 10 min max
        finally:
            self._broadcast = orig_broadcast  # type: ignore

    def submit_prompt(self, prompt: str) -> bool:
        """Submit a prompt or slash command. Returns False if agent is busy."""
        # Handle slash commands locally (don't send to LLM)
        if prompt.startswith("/") and not self._pending_attachments:
            return self._handle_slash(prompt)

        if self._busy.is_set():
            self._broadcast(ChatEvent("error", {"message": "Agent is busy"}))
            return False

        self.last_active = time.monotonic()
        attachment_context, attachment_images = self._consume_attachment_context()

        # When images are attached and an auxiliary vision model is configured,
        # pre-process images through the auxiliary model and inject text descriptions.
        # This keeps text-only primary models (e.g. DeepSeek) working with images.
        if attachment_images:
            vision_text = self._run_auxiliary_vision(attachment_images)
            if vision_text:
                attachment_context = (
                    (attachment_context or "") +
                    "\n\n[辅助视觉模型对上传图片的分析描述]\n" +
                    vision_text
                )
                # Don't send raw images to primary model — text description suffices
                attachment_images = []

        agent_prompt = f"{attachment_context}\n\n用户问题：{prompt}" if attachment_context else prompt
        if attachment_images:
            import runtime
            ctx = runtime.get_session_ctx(self.session_id)
            ctx.pending_image_parts.extend(attachment_images)
        # Clear event buffer for fresh turn — don't replay stale events
        with self._sub_lock:
            self._event_buffer.clear()
        self._append_msg({"role": "user", "content": prompt})
        self._broadcast(ChatEvent("status", {"state": "running"}))

        def _run():
            self._busy.set()
            try:
                self._run_agent(agent_prompt)
            except Exception as exc:
                self._broadcast(ChatEvent("error", {
                    "message": str(exc),
                    "traceback": traceback.format_exc(),
                }))
            finally:
                self._busy.clear()
                self._broadcast(ChatEvent("status", {"state": "idle"}))

        self._agent_thread = threading.Thread(target=_run, daemon=True)
        self._agent_thread.start()
        return True

    def _run_auxiliary_vision(self, images: list[dict]) -> str:
        """Run vision analysis on attached images using the auxiliary model.

        Returns a text description of all images, or empty string on failure.
        The auxiliary model must be vision-capable (GPT-4o, Gemini Flash, etc.).
        """
        try:
            import auxiliary
        except ImportError:
            return ""

        try:
            aux_model = auxiliary.get_auxiliary_model(self.config)
        except Exception:
            return ""

        # If auxiliary model is same as primary, skip — avoid double call
        primary = self.config.get("model", "")
        if aux_model == primary:
            return ""

        # Build a user message with images for the auxiliary model
        image_blocks = []
        for img in images:
            mime = str(img.get("mime") or "image/png")
            data = str(img.get("data") or "")
            if not data:
                continue
            image_blocks.append({"data": data, "mime": mime})

        if not image_blocks:
            return ""

        system_prompt = (
            "You are an image analysis assistant. Describe each image in detail: "
            "what objects, people, text, colors, layout, and context you see. "
            "Include all visible text verbatim. Be thorough but concise."
        )

        user_msg: dict = {"role": "user", "content": "Please describe these images in detail."}
        if image_blocks:
            user_msg["images"] = image_blocks

        try:
            text = auxiliary.stream_auxiliary(
                system=system_prompt,
                messages=[user_msg],
                config=self.config,
            )
            return text.strip() if text else ""
        except Exception:
            return ""

    def _handle_slash(self, line: str) -> bool:
        """Handle /commands locally, capture stdout, broadcast as system message.

        Some commands return sentinel tuples that require follow-up agent runs
        (e.g. __brainstorm__, __worker__, __plan__, __ssj_cmd__).  These are
        executed on a background thread exactly like a regular prompt.
        """
        import io
        import re as _re
        self.last_active = time.monotonic()

        self._append_msg({"role": "user", "content": line})

        # Parse command and args
        cmd_parts = line[1:].split(None, 1)
        cmd_name = cmd_parts[0].lower() if cmd_parts else ""
        cmd_args = cmd_parts[1].strip() if len(cmd_parts) > 1 else ""

        if self.config.get("rich_business_mode"):
            allowed_commands = {"help", "clear", "status", "model", "config"}
            if cmd_name not in allowed_commands:
                output = (
                    f"RICH 业务模式不支持 /{cmd_name} 命令。"
                    "请直接用自然语言询问投资组合、持仓、交易、定投、页面导航或任务/工作流。"
                )
                self._append_msg({"role": "assistant", "content": output})
                self._broadcast(ChatEvent("command_result", {
                    "command": line,
                    "output": output,
                }))
                return True

        # /ssj with no args → show interactive menu
        if cmd_name == "ssj" and not cmd_args:
            self._broadcast(ChatEvent("interactive_menu", {
                "command": line,
                "menu": "ssj",
                "items": [
                    {"key":"1",  "icon":"bulb",    "label":"Brainstorm",    "cmd":"/brainstorm"},
                    {"key":"2",  "icon":"clipboard","label":"Show TODO",     "cmd":"/ssj todo"},
                    {"key":"3",  "icon":"worker",   "label":"Worker",        "cmd":"/worker"},
                    {"key":"4",  "icon":"brain",    "label":"Debate File",   "cmd":"/ssj debate"},
                    {"key":"5",  "icon":"sparkle",  "label":"Propose Improvement","cmd":"/ssj propose"},
                    {"key":"6",  "icon":"search",   "label":"Review File",   "cmd":"/ssj review"},
                    {"key":"7",  "icon":"book",     "label":"Generate README","cmd":"/ssj readme"},
                    {"key":"8",  "icon":"chat",     "label":"AI Commit Msg", "cmd":"/ssj commit"},
                    {"key":"9",  "icon":"test",     "label":"Scan Git Diff", "cmd":"/ssj scan"},
                    {"key":"10", "icon":"note",     "label":"Promote to Tasks","cmd":"/ssj promote"},
                    {"key":"13", "icon":"monitor",  "label":"Monitor",       "cmd":"/monitor"},
                    {"key":"15", "icon":"robot",    "label":"Autonomous Agent","cmd":"/agent"},
                ],
            }))
            return True

        # /ssj <subcommand> → map to direct actions (skip the interactive menu)
        _SSJ_DIRECT = {
            "debate":  ("__ssj_query__", "Act as a panel of 3 expert engineers. Each gives 2-3 critical insights on the codebase. Be specific and constructive."),
            "propose": ("__ssj_query__", "Analyze the codebase and propose 3 high-impact improvements with code examples. Focus on correctness, performance, or maintainability."),
            "review":  ("__ssj_query__", "Give a quick code review: identify bugs, code smells, or missing edge cases. Be concise."),
            "readme":  ("__ssj_query__", "Generate a comprehensive README.md for this project. Include: project description, features, installation, usage examples, and contributing guidelines."),
            "commit":  ("__ssj_query__", "Review the git diff (git diff HEAD) and suggest a concise, descriptive commit message following conventional commits format. Also list files changed."),
            "scan":    ("__ssj_query__", "Run git diff HEAD and analyze the changes. Summarize what was changed, why it might have been changed, and flag any potential issues or regressions."),
            "todo":    None,  # handled below
        }
        if cmd_name == "ssj" and cmd_args.split()[0].lower() in _SSJ_DIRECT:
            sub = cmd_args.split()[0].lower()
            extra_args = cmd_args[len(sub):].strip()
            if sub == "todo":
                pass  # fall through to normal handler
            else:
                action = _SSJ_DIRECT[sub]
                prompt = action[1]
                if extra_args:
                    prompt += f" Focus on: {extra_args}"
                # Run as agent query
                self._broadcast(ChatEvent("status", {"state": "running"}))
                self._broadcast(ChatEvent("command_result", {
                    "command": line,
                    "output": f"Running SSJ {sub}...",
                }))

                def _run_ssj():
                    self._busy.set()
                    import runtime
                    ctx = runtime.get_session_ctx(self.session_id)
                    ctx.in_web_turn = True
                    try:
                        self._run_agent(prompt)
                    except Exception as exc:
                        self._broadcast(ChatEvent("error",
                                                  {"message": str(exc)}))
                    finally:
                        ctx.in_web_turn = False
                        self._busy.clear()
                        self._broadcast(ChatEvent("status",
                                                  {"state": "idle"}))

                self._agent_thread = threading.Thread(target=_run_ssj,
                                                      daemon=True)
                self._agent_thread.start()
                return True

        # /brainstorm with no topic → ask for topic via input_request event
        if cmd_name == "brainstorm" and not cmd_args:
            self._broadcast(ChatEvent("input_request", {
                "command": "/brainstorm",
                "prompt": "Brainstorm topic (Enter for general):",
                "placeholder": "e.g. improve test coverage, refactor auth...",
                "default_cmd": "/brainstorm general project improvement",
            }))
            return True

        # Long-running commands — run on background thread with live events.
        # These call providers.stream() internally and take minutes.
        # We redirect stdout so their print() output streams to the browser.
        _LONG_RUNNING = {"brainstorm", "worker", "agent", "plan"}
        if cmd_name in _LONG_RUNNING:
            self._broadcast(ChatEvent("status", {"state": "running"}))
            session_ref = self  # capture for closure

            # Thread-local stdout wrapper: intercepts print() calls from
            # the command handler and broadcasts them as text_chunk events.
            # Uses threading.current_thread() check to avoid affecting other threads.
            _target_thread_id = [None]  # set inside the thread

            class _ThreadLocalStdout:
                """Only intercepts writes from the target thread."""
                def __init__(self, broadcast_fn, real):
                    self._broadcast = broadcast_fn
                    self._real = real
                def write(self, s):
                    if not s:
                        return
                    if threading.current_thread().ident == _target_thread_id[0]:
                        import re as _re2
                        clean = _re2.sub(r'\x1b\[[0-9;]*m', '', s)
                        if clean.strip():
                            safe = session_ref._filter_text(clean)
                            self._broadcast(ChatEvent("text_chunk",
                                                      {"text": safe}))
                    else:
                        self._real.write(s)
                def flush(self):
                    self._real.flush()
                # Forward attributes to real stdout for compatibility
                def fileno(self):
                    return self._real.fileno()
                @property
                def encoding(self):
                    return getattr(self._real, 'encoding', 'utf-8')

            def _run_long():
                _target_thread_id[0] = threading.current_thread().ident
                self._busy.set()
                import runtime
                ctx = runtime.get_session_ctx(self.session_id)
                ctx.in_web_turn = True
                wrapper = _ThreadLocalStdout(session_ref._broadcast,
                                            sys.stdout)
                old_out, old_err = sys.stdout, sys.stderr
                sys.stdout = wrapper
                sys.stderr = wrapper
                try:
                    result = _web_handle_slash(line, self._agent_state,
                                              self.config)
                    if isinstance(result, tuple):
                        self._process_sentinel(result)
                    elif result is True:
                        self._broadcast(ChatEvent("command_result", {
                            "command": line,
                            "output": "(done)",
                        }))
                except Exception as exc:
                    self._broadcast(ChatEvent("error",
                                              {"message": str(exc)}))
                finally:
                    sys.stdout = old_out
                    sys.stderr = old_err
                    ctx.in_web_turn = False
                    self._busy.clear()
                    self._broadcast(ChatEvent("status", {"state": "idle"}))

            self._agent_thread = threading.Thread(target=_run_long,
                                                  daemon=True)
            self._agent_thread.start()
            return True

        # Quick commands — capture stdout synchronously
        capture = io.StringIO()
        old_stdout, old_stderr = sys.stdout, sys.stderr
        try:
            sys.stdout = capture
            sys.stderr = capture
            result = _web_handle_slash(line, self._agent_state, self.config)
        except Exception as exc:
            self._broadcast(ChatEvent("error", {"message": str(exc)}))
            return True
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr

        output = capture.getvalue().strip()
        output = _re.sub(r'\x1b\[[0-9;]*m', '', output)

        if output:
            self._append_msg({"role": "assistant", "content": output})
            self._broadcast(ChatEvent("command_result", {
                "command": line, "output": output,
            }))

        # Handle sentinel tuples from quick commands (unlikely but safe)
        if isinstance(result, tuple):
            self._broadcast(ChatEvent("status", {"state": "running"}))

            def _run_sentinel():
                self._busy.set()
                try:
                    self._process_sentinel(result)
                except Exception as exc:
                    self._broadcast(ChatEvent("error",
                                              {"message": str(exc)}))
                finally:
                    self._busy.clear()
                    self._broadcast(ChatEvent("status", {"state": "idle"}))

            self._agent_thread = threading.Thread(target=_run_sentinel,
                                                  daemon=True)
            self._agent_thread.start()
            return True

        if not output and result is True:
            self._broadcast(ChatEvent("command_result", {
                "command": line, "output": "(done)",
            }))

        return True

    def _process_sentinel(self, result: tuple):
        """Execute the multi-step workflow described by a sentinel tuple."""
        sentinel = result[0]

        if sentinel == "__brainstorm__":
            _, brain_prompt, brain_out_file = result
            self._broadcast(ChatEvent("command_result", {
                "command": "/brainstorm",
                "output": "Starting multi-persona brainstorm...",
            }))
            self._run_agent(brain_prompt)
            # Generate todo list from synthesis
            from pathlib import Path
            todo_path = str(Path(brain_out_file).parent / "todo_list.txt")
            self._run_agent(
                f"Based on the Master Plan you just synthesized, generate a "
                f"todo list file at {todo_path}. Format: one task per line, "
                f"each starting with '- [ ] '. Order by priority. Include ALL "
                f"actionable items from the plan. Use the Write tool to create "
                f"the file. Do NOT explain, just write the file now."
            )

        elif sentinel == "__worker__":
            _, worker_tasks = result
            total = len(worker_tasks)
            for i, (line_idx, task_text, prompt) in enumerate(worker_tasks):
                self._broadcast(ChatEvent("command_result", {
                    "command": f"/worker ({i+1}/{total})",
                    "output": task_text,
                }))
                self._run_agent(prompt)

        elif sentinel == "__plan__":
            _, plan_desc = result
            self._broadcast(ChatEvent("command_result", {
                "command": "/plan",
                "output": f"Entering plan mode: {plan_desc}",
            }))
            self._run_agent(
                f"Please analyze the codebase and create a detailed "
                f"implementation plan for: {plan_desc}"
            )

        elif sentinel == "__ssj_cmd__":
            # SSJ delegates to another slash command
            _, cmd_name, cmd_args = result
            inner_line = f"/{cmd_name} {cmd_args}".strip()
            self._broadcast(ChatEvent("command_result", {
                "command": "/ssj",
                "output": f"Executing: {inner_line}",
            }))
            # Re-enter slash handling for the delegated command
            self._handle_slash_inner(inner_line)

        elif sentinel in ("__ssj_query__", "__ssj_debate__",
                          "__ssj_passthrough__", "__ssj_promote_worker__"):
            # These carry a prompt to run through the agent
            prompt = result[1] if len(result) > 1 else ""
            if prompt:
                self._run_agent(prompt)

        elif sentinel == "__image__":
            self._broadcast(ChatEvent("command_result", {
                "command": "/image",
                "output": "Image/vision: paste an image URL or use the terminal for clipboard support.",
            }))

        elif sentinel == "__voice__":
            self._broadcast(ChatEvent("command_result", {
                "command": "/voice",
                "output": "Voice input requires the terminal (microphone access).",
            }))

        else:
            # Unknown sentinel — try to extract a prompt if it has one
            if len(result) > 1 and isinstance(result[1], str) and result[1]:
                self._run_agent(result[1])
            else:
                self._broadcast(ChatEvent("command_result", {
                    "command": str(result[0]),
                    "output": "This feature may require the terminal for full support.",
                }))

    def _handle_slash_inner(self, line: str):
        """Re-entrant slash handling for SSJ delegation."""
        import io
        import re as _re

        capture = io.StringIO()
        old_stdout, old_stderr = sys.stdout, sys.stderr
        try:
            sys.stdout = capture
            sys.stderr = capture
            result = _web_handle_slash(line, self._agent_state, self.config)
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr

        output = _re.sub(r'\x1b\[[0-9;]*m', '', capture.getvalue().strip())
        if output:
            self._append_msg({"role": "assistant", "content": output})
            self._broadcast(ChatEvent("command_result", {
                "command": line, "output": output,
            }))

        if isinstance(result, tuple):
            self._process_sentinel(result)

    def _run_agent(self, prompt: str):
        """Iterate agent.run() generator, broadcast events."""
        from agent import (run, TextChunk, ThinkingChunk, ToolStart,
                           ToolEnd, TurnDone, PermissionRequest)
        from context import build_system_prompt
        import runtime

        ctx = runtime.get_session_ctx(self.session_id)
        ctx.in_web_turn = True
        system_prompt = build_system_prompt(self.config)
        if self.config.get("rich_business_mode"):
            system_prompt += """

# RICH Business Agent Mode

你是 RICH 财富管理系统内的业务助手，不是代码助手，也不是通用 AI 助手。

## 核心规则（必须遵守）

1. **绝不暴露实现细节**：你的回答中不能出现服务器路径、代码文件名、函数名、数据库表名、端口号、IP 地址、技术栈名称等实现细节。用户不需要知道这些，也绝不能让他们知道。
2. **只使用可用工具**：你只能通过当前可用的 RICH 业务工具帮助用户。不要尝试用其他方式获取信息。
3. **不要猜测或编造路径**：导航时只使用 navigate 工具，不要自己猜测 URL 或告诉用户"试试 /xxx 路径"。你不知道系统有哪些页面，让工具告诉你。
4. **文件仅限用户上传**：你只能阅读用户主动上传的附件内容。不要说"让我读取这个文件"、"我先看看代码"之类的话。
5. **附件优先且不可替代**：用户说"这个附件"、"这个表格"、"这个交割单"时，只能基于本轮附件内容回答；如果附件解析失败或没有可读表格内容，必须明确说没法读取该附件，不要调用交易查询工具、不要用系统已有交易记录冒充附件内容。
6. **禁止执行命令**：你不能执行终端命令、读取服务器文件、扫描目录、访问数据库。

## 可用业务工具
- Skill / SkillList — 使用项目预置业务 skill，例如上传表格导入流程
- navigate / open_symbol_chart — 打开系统页面或指定标的 K 线图
- get_portfolios / get_portfolio_detail / create_portfolio / update_portfolio / delete_portfolio — 投资组合查询和增删改（写/删需确认）
- get_asset_types / get_asset_type_preset / create_asset_type / update_asset_type / delete_asset_type — 资产类型管理；交易和持仓创建前必须先有资产类型
- get_asset_positions / get_asset_allocation / get_portfolio_assets / get_assets_summary / add_portfolio_asset / update_portfolio_asset / update_portfolio_asset_price / delete_portfolio_asset — 持仓资产查询和增删改（写/删需确认）
- get_recent_transactions / get_transactions / create_transaction / update_transaction / delete_transaction / update_transaction_status / get_transaction_statistics — 交易记录查询、快速记录、修改、删除和统计（写/删需确认）
- preview_business_import / execute_business_import — 预览并执行用户上传表格中的资产类型和交易记录导入（执行需确认）
- get_asset_groups / get_asset_group_detail / create_asset_group / update_asset_group / delete_asset_group / add_asset_group_member / update_asset_group_member / remove_asset_group_member / validate_asset_group_weights / get_group_value — 资产配置分组管理（写/删需确认）
- analyze_portfolio_risk / analyze_portfolio_performance — 分析组合风险、集中度、偏离和收益表现
- calculate_rebalance_plan — 预览再平衡方案，不写入数据
- execute_rebalance / force_recalculate_portfolio / update_portfolio_weights — 执行再平衡、强制重算或更新目标权重（需确认）
- get_dca_plans / get_dca_plan_detail / delete_dca_plan / preview_dca_allocation / get_pending_dca_plans / get_dca_execution_history / get_dca_statistics — 定投计划查询、删除、预览、历史和统计
- create_dca_plan / update_dca_plan / toggle_dca_plan / execute_dca_plan / run_due_dca_plans — 创建、更新、启停、执行或扫描到期定投（需确认）
- search_market_symbols / list_market_symbols / get_kline_history / analyze_kline / query_valuation_data / query_factors — 标的搜索、K 线、估值和因子查询分析

## 工具选择规则
- 用户问业务数据时，先调用对应查询或预览工具，不要猜测。
- 用户要快速记录交易时，先确认资产类型 code 是否存在；不知道时先 get_asset_types，不存在先 create_asset_type，再 create_transaction。
- 用户要删除/修改交易记录时，先用 get_transactions 定位记录，明确操作后再 delete_transaction 或 update_transaction。
- 用户上传 Excel/CSV 并要求导入资产类型或交易记录时，先从附件内容解析出 asset_types 和 transactions，调用 preview_business_import 展示将创建的类型、将导入的交易、错误和疑似重复；用户确认后调用 execute_business_import。
- 导入交易前必须保证资产类型存在；preview_business_import/execute_business_import 可自动补全缺失资产类型。
- 用户要添加持仓资产时，先确认资产类型存在；不存在先 create_asset_type，再 add_portfolio_asset。
- 用户要打开某个标的 K 线时调用 open_symbol_chart；要分析 K 线时调用 analyze_kline，不要只跳转页面。
- 用户要求“制定/创建定投计划”，参数足够时调用 create_dca_plan；只想看分配时调用 preview_dca_allocation。
- 用户要求“再平衡方案”时调用 calculate_rebalance_plan；明确“执行/重算/应用”时调用 execute_rebalance 或 force_recalculate_portfolio。
- 用户要求“数据更新/同步/巡检/因子计算/市值重算”或后台工作流时，告知这些由管理员在系统设置/数据管理中手动管理，或由系统自动任务处理，Agent 暂不执行。

## 回答风格
- 用表格和简洁的中文回复，面向投资业务用户
- 不要提"我调用了 XX 工具"——用户不需要知道你在后台用什么工具
- 如果工具返回错误或空数据，直接告诉用户当前无法获取，不要解释技术原因
- 绝对不要输出任何文件路径、URL、代码片段或服务器信息
"""

        text_chunks: list[str] = []
        tool_calls: list[dict] = []

        # Do NOT wire RuntimeContext callbacks — we broadcast from the
        # generator loop below.  Wiring ctx.on_text_chunk etc. would cause
        # duplicate events because the REPL's run_query() also fires them
        # for every yielded event.  Single event source = generator loop only.
        ctx.on_text_chunk = None
        ctx.on_tool_start = None
        ctx.on_tool_end = None

        try:
            for event in run(prompt, self._agent_state, self.config,
                             system_prompt):
                if isinstance(event, TextChunk):
                    filtered = self._filter_text(event.text)
                    text_chunks.append(filtered)
                    self._broadcast(ChatEvent("text_chunk",
                                              {"text": filtered}))

                elif isinstance(event, ThinkingChunk):
                    self._broadcast(ChatEvent("thinking_chunk",
                                              {"text": self._filter_text(event.text)}))

                elif isinstance(event, ToolStart):
                    tool_calls.append({
                        "name": event.name,
                        "inputs": event.inputs,
                        "status": "running",
                    })
                    self._broadcast(ChatEvent("tool_start", {
                        "name": event.name,
                        "inputs": event.inputs,
                    }))

                elif isinstance(event, PermissionRequest):
                    self._broadcast(ChatEvent("permission_request", {
                        "description": event.description,
                    }))
                    # Block until browser responds
                    evt = threading.Event()
                    ctx.web_input_event = evt
                    try:
                        if evt.wait(timeout=300):
                            val = ctx.web_input_value.strip().lower()
                            event.granted = val in ("y", "yes", "true", "1")
                        else:
                            event.granted = False
                            self._broadcast(ChatEvent("error", {
                                "message": "Permission request timed out (5 min)",
                            }))
                    finally:
                        # Always clean up — prevents dangling event objects
                        ctx.web_input_event = None
                        ctx.web_input_value = ""
                    self._broadcast(ChatEvent("permission_response", {
                        "granted": event.granted,
                    }))

                elif isinstance(event, ToolEnd):
                    for tc in reversed(tool_calls):
                        if tc["name"] == event.name and tc["status"] == "running":
                            tc["status"] = "done" if event.permitted else "denied"
                            tc["result"] = event.result[:2000] if event.result else ""
                            break
                    self._broadcast(ChatEvent("tool_end", {
                        "name": event.name,
                        "result": event.result[:2000] if event.result else "",
                        "permitted": event.permitted,
                    }))

                elif isinstance(event, TurnDone):
                    self._broadcast(ChatEvent("turn_done", {
                        "input_tokens": event.input_tokens,
                        "output_tokens": event.output_tokens,
                    }))

            # Store assistant response in history
            final_text = "".join(text_chunks)
            msg: dict = {"role": "assistant", "content": final_text}
            if tool_calls:
                msg["tool_calls"] = tool_calls
            self._append_msg(msg)

        except Exception as exc:
            self._broadcast(ChatEvent("error", {"message": str(exc)}))
        finally:
            ctx.on_text_chunk = None
            ctx.on_tool_start = None
            ctx.on_tool_end = None
            ctx.in_web_turn = False

    # ── Permission approval ────────────────────────────────────────────

    def approve_permission(self, granted: bool):
        """Respond to a pending PermissionRequest."""
        import runtime
        ctx = runtime.get_session_ctx(self.session_id)
        evt = ctx.web_input_event
        if evt:
            ctx.web_input_value = "y" if granted else "n"
            evt.set()

    # ── Introspection ──────────────────────────────────────────────────

    def _append_msg(self, msg: dict):
        with self._msg_lock:
            self.messages.append(msg)
        # Persist to DB (best-effort; don't break streaming on DB failure)
        try:
            from web import db as _db
            _db.repo.append_message(
                self.session_id,
                msg.get("role", "system"),
                msg.get("content", "") or "",
                msg.get("tool_calls"),
            )
            # Keep in-memory title in sync with auto-titling in repo
            sess = _db.repo.get_session(self.session_id, self.user_id)
            if sess and sess["title"] != self.title:
                self.title = sess["title"]
        except Exception as exc:  # noqa: BLE001
            from web.logging_setup import get_logger
            get_logger("api").exception("message persist failed",
                                         extra={"session_id": self.session_id,
                                                "err": str(exc)})

    def get_messages(self) -> list[dict]:
        with self._msg_lock:
            return list(self.messages)

    def get_safe_config(self) -> dict:
        result = {k: self.config.get(k) for k in _SAFE_CONFIG_KEYS
                  if k in self.config}
        # Show which providers have API keys configured (without revealing them)
        result["api_keys_configured"] = {
            provider: bool(self.config.get(cfg_key) or
                          os.environ.get(cfg_key.upper(), ""))
            for provider, cfg_key in _API_KEY_CONFIG_MAP.items()
        }
        result["custom_base_url"] = self.config.get("custom_base_url", "")
        result["ollama_base_url"] = self.config.get("ollama_base_url",
                                                     "http://localhost:11434")
        return result

    def update_config(self, updates: dict) -> dict:
        for k, v in updates.items():
            if k in _WRITABLE_CONFIG_KEYS:
                self.config[k] = v
        # Persist non-secret config keys to DB (secrets stay session-only)
        try:
            from web import db as _db
            _db.repo.upsert_session(
                self.session_id, self.user_id,
                title=self.title,
                config={k: v for k, v in self.config.items()
                        if k in _SAFE_CONFIG_KEYS},
            )
        except Exception as exc:  # noqa: BLE001
            from web.logging_setup import get_logger
            get_logger("api").exception("config persist failed",
                                         extra={"session_id": self.session_id,
                                                "err": str(exc)})
        return self.get_safe_config()

    def is_idle(self) -> bool:
        return not self._busy.is_set()

    def is_stale(self) -> bool:
        return (time.monotonic() - self.last_active) > _IDLE_TIMEOUT

    # ── Cleanup ────────────────────────────────────────────────────────

    def cleanup(self):
        import runtime
        runtime.release_session_ctx(self.session_id)


# ── Session registry ───────────────────────────────────────────────────────

_chat_sessions: dict[str, ChatSession] = {}
_chat_lock = threading.Lock()


def create_chat_session(base_config: dict, user_id: int) -> ChatSession:
    session = ChatSession(base_config, user_id=user_id)
    with _chat_lock:
        _chat_sessions[session.session_id] = session
    return session


def get_chat_session(sid: str,
                    user_id: Optional[int] = None,
                    base_config: Optional[dict] = None) -> Optional[ChatSession]:
    """Return a live ChatSession, hydrating from DB if necessary.

    If the session isn't in the in-memory cache but exists in the DB (and is
    owned by `user_id`), it's lazily rehydrated so restarts don't lose state.
    `user_id` is required for DB hydration; pass None to skip hydration and
    only look in memory (used by internal callers that already validated).
    """
    with _chat_lock:
        sess = _chat_sessions.get(sid)
        if sess:
            # Enforce ownership even on cache hits — otherwise users could
            # read each other's sessions whenever the cache is warm.
            if user_id is not None and sess.user_id != user_id:
                return None
            return sess
    if user_id is None or base_config is None:
        return None
    # Try to hydrate from DB
    try:
        from web import db as _db
        row = _db.repo.get_session(sid, user_id)
    except Exception:  # noqa: BLE001
        return None
    if not row:
        return None
    session = ChatSession(base_config, user_id=user_id, session_id=sid)
    with _chat_lock:
        # Guard against a race where another thread hydrated concurrently.
        existing = _chat_sessions.get(sid)
        if existing:
            return existing
        _chat_sessions[sid] = session
    return session


def list_chat_sessions(user_id: int) -> list[dict]:
    """List this user's sessions (DB is the source of truth, not memory)."""
    try:
        from web import db as _db
        rows = _db.repo.list_sessions(user_id)
    except Exception as exc:  # noqa: BLE001
        from web.logging_setup import get_logger
        get_logger("api").exception("list_sessions failed",
                                     extra={"user_id": user_id,
                                            "err": str(exc)})
        rows = []
    busy_ids = set()
    with _chat_lock:
        for sid, s in _chat_sessions.items():
            if s._busy.is_set():
                busy_ids.add(sid)
    return [{**r, "busy": r["id"] in busy_ids} for r in rows]


def remove_chat_session(sid: str, user_id: int) -> bool:
    """Remove session from DB and in-memory cache. Returns True if removed."""
    try:
        from web import db as _db
        deleted = _db.repo.delete_session(sid, user_id)
    except Exception as exc:  # noqa: BLE001
        from web.logging_setup import get_logger
        get_logger("api").exception("delete_session failed",
                                     extra={"session_id": sid,
                                            "user_id": user_id,
                                            "err": str(exc)})
        deleted = False
    with _chat_lock:
        session = _chat_sessions.pop(sid, None)
    if session:
        session.cleanup()
    return deleted


def list_folders(user_id: int) -> list[dict]:
    from web import db as _db
    return _db.repo.list_folders(user_id)


def create_folder(user_id: int, name: str) -> Optional[dict]:
    from web import db as _db
    return _db.repo.create_folder(user_id, name)


def rename_folder(folder_id: int, user_id: int, name: str) -> bool:
    from web import db as _db
    return _db.repo.rename_folder(folder_id, user_id, name)


def remove_folder(folder_id: int, user_id: int) -> bool:
    from web import db as _db
    return _db.repo.delete_folder(folder_id, user_id)


def move_session_to_folder(sid: str, user_id: int,
                            folder_id: Optional[int]) -> bool:
    from web import db as _db
    return _db.repo.move_session_to_folder(sid, user_id, folder_id)


def batch_remove_chat_sessions(sids: list, user_id: int) -> dict:
    """Delete multiple sessions for a user. Cross-user IDs are silently
    skipped (delete_session enforces ownership). Returns counts."""
    deleted = 0
    failed: list[str] = []
    for sid in sids:
        try:
            if remove_chat_session(sid, user_id):
                deleted += 1
            else:
                failed.append(sid)
        except Exception:  # noqa: BLE001
            failed.append(sid)
    return {"deleted": deleted, "failed": failed, "requested": len(sids)}


def batch_export_chat_sessions_markdown(sids: list,
                                         user_id: int) -> Optional[str]:
    """Combine multiple sessions into a single markdown document. Returns
    None when no requested session belongs to the user."""
    parts: list[str] = []
    rendered = 0
    for sid in sids:
        md = export_chat_session_markdown(sid, user_id)
        if md is None:
            continue
        rendered += 1
        if parts:
            parts.append("\n\n---\n\n")
        parts.append(md)
    if rendered == 0:
        return None
    import datetime as _dt
    header = (
        f"# Chat Export — {rendered} session"
        f"{'s' if rendered != 1 else ''}\n\n"
        f"- Exported: {_dt.datetime.now():%Y-%m-%d %H:%M}\n"
        f"- User ID: {user_id}\n\n---\n\n"
    )
    return header + "".join(parts)


def rename_chat_session(sid: str, user_id: int, title: str) -> bool:
    try:
        from web import db as _db
        ok = _db.repo.rename_session(sid, user_id, title)
    except Exception:  # noqa: BLE001
        return False
    if ok:
        with _chat_lock:
            s = _chat_sessions.get(sid)
            if s:
                s.title = title.strip()[:200] or "Untitled"
    return ok


def export_chat_session_markdown(sid: str, user_id: int) -> Optional[str]:
    """Render a session's messages as Markdown. Returns None if not found."""
    try:
        from web import db as _db
        meta = _db.repo.get_session(sid, user_id)
        if not meta:
            return None
        msgs = _db.repo.get_messages(sid)
    except Exception:  # noqa: BLE001
        return None
    import datetime as _dt
    lines: list[str] = []
    lines.append(f"# {meta['title']}")
    lines.append("")
    lines.append(f"- Session ID: `{sid}`")
    lines.append(f"- Created: {_dt.datetime.fromtimestamp(meta['created_at']):%Y-%m-%d %H:%M}")
    lines.append(f"- Messages: {len(msgs)}")
    lines.append("")
    lines.append("---")
    lines.append("")
    for m in msgs:
        role = m.get("role", "?")
        when = _dt.datetime.fromtimestamp(m.get("created_at", 0)).strftime("%H:%M:%S")
        lines.append(f"## {role.title()} · {when}")
        lines.append("")
        lines.append(m.get("content", "") or "_(no content)_")
        if m.get("tool_calls"):
            lines.append("")
            lines.append("<details><summary>Tool calls</summary>")
            lines.append("")
            for tc in m["tool_calls"]:
                lines.append(f"- **{tc.get('name','?')}** "
                             f"(status: {tc.get('status','?')})")
                if tc.get("inputs"):
                    import json as _j
                    lines.append("  ```json")
                    lines.append("  " + _j.dumps(tc["inputs"], indent=2)
                                 .replace("\n", "\n  "))
                    lines.append("  ```")
            lines.append("")
            lines.append("</details>")
        lines.append("")
    return "\n".join(lines)


def get_available_models() -> list[dict]:
    """Return all providers and their models for the UI model picker."""
    try:
        from providers import PROVIDERS
    except ImportError:
        return []
    result = []
    for name, info in PROVIDERS.items():
        result.append({
            "provider": name,
            "models": list(info.get("models", [])),
            "context_limit": info.get("context_limit", 128000),
            "needs_api_key": info.get("api_key_env") is not None,
            "has_api_key": bool(
                os.environ.get(info.get("api_key_env") or "", "") or
                info.get("api_key", "")
            ),
        })
    return result


def reap_stale_chat_sessions():
    """Called periodically by server.py's reaper thread.

    `remove_chat_session` requires the owning user_id for ownership-check
    parity with the per-user DELETE endpoint, so we capture it from the
    cached ChatSession object — collecting `(sid, user_id)` pairs under the
    lock and applying outside it (remove_chat_session re-acquires).
    """
    stale: list[tuple[str, int]] = []
    with _chat_lock:
        for sid, session in _chat_sessions.items():
            if session.is_stale() and session.is_idle():
                stale.append((sid, session.user_id))
    for sid, user_id in stale:
        remove_chat_session(sid, user_id)
